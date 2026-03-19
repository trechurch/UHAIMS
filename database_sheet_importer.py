"""
database_sheet_importer.py
One-shot import of the DATABASE sheet from 2026_UH__2_.xlsx
Run from repo root: streamlit run database_sheet_importer.py

Imports 402 manually curated concessions items with overrides locked.
Then imports the combined count sheet to add quantities.
Reports final totals vs MyOrders target.
"""

import re
import openpyxl
import streamlit as st
import pandas as pd
from datetime import datetime
from pathlib import Path

st.set_page_config(page_title="DB Sheet Importer", page_icon="📋", layout="wide")

# ── Connection ────────────────────────────────────────────────────────────────
@st.cache_resource
def get_db():
    from database import InventoryDatabase
    return InventoryDatabase()

# ── Helpers ───────────────────────────────────────────────────────────────────
def build_key(name, pack):
    n = str(name or "").strip().upper()
    p = str(pack or "").strip().upper()
    if not n:
        return None
    return f"{n}||{p}" if p else f"{n}||CASE"

def safe_float(v, default=0.0):
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default

# ── Step 1: Load DATABASE sheet ───────────────────────────────────────────────
def load_database_sheet(filepath):
    wb  = openpyxl.load_workbook(filepath, data_only=True)
    ws  = wb['DATABASE']
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        item_name  = str(row[0]).strip()
        pack_type  = str(row[1] or "").strip()
        cost       = safe_float(row[2])
        per        = str(row[3] or "Case").strip()
        conv_ratio = safe_float(row[4], 1.0)
        units      = str(row[5] or "Each").strip()
        vendor     = str(row[6] or "").strip()
        item_num   = str(row[7] or "").strip()
        mog        = str(row[8] or "").strip()
        brand      = str(row[10] or "").strip()
        yield_val  = safe_float(row[13], 1.0)

        key = build_key(item_name, pack_type)
        if not key:
            continue

        items.append({
            "key":                  key,
            "description":          item_name.upper(),
            "pack_type":            pack_type,
            "cost":                 cost,
            "per":                  per,
            "conv_ratio":           conv_ratio,
            "unit":                 units,
            "vendor":               vendor,
            "item_number":          item_num,
            "mog":                  mog,
            "brand":                brand,
            "yield":                yield_val,
            "override_conv_ratio":  conv_ratio if conv_ratio != 1.0 else None,
            "override_yield":       yield_val  if yield_val  != 1.0 else None,
            "override_pack_type":   pack_type  if pack_type  else None,
            "override_vendor":      vendor     if vendor     else None,
            "status_tag":           "📋 DATABASE Import",
            "record_status":        "active",
            "cost_center":          "57231",
            "quantity_on_hand":     0,
            "is_chargeable":        True,
            "last_updated":         datetime.utcnow(),
            "created_date":         datetime.utcnow(),
        })
    return items

# ── Step 2: Load combined count sheet ────────────────────────────────────────
def load_count_sheet(filepath):
    wb  = openpyxl.load_workbook(filepath, data_only=True)
    ws  = wb['CurrentInventory']
    counts = {}  # key -> total_each_qty
    prices = {}  # key -> each_price

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        desc = row[2]
        if not desc or desc == 'Item Description':
            continue

        count_raw = str(row[3] or "0")
        pack_type = str(row[4] or "").strip()
        price_raw = str(row[5] or "0")

        # Parse count: "4.00 Case/14.00 Each"
        case_qty = 0.0
        each_qty = 0.0
        cm = re.match(r'([\d.]+)\s*Case.*?/([\d.]+)\s*Each', count_raw, re.I)
        if cm:
            case_qty = float(cm.group(1))
            each_qty = float(cm.group(2))
        else:
            try:
                each_qty = float(re.sub(r'[^\d.]', '', count_raw.split('/')[0]))
            except Exception:
                pass

        # Parse price: "$30.46/$1.27" → each price is second
        each_price = 0.0
        pm = re.match(r'\$([\d.]+)/\$([\d.]+)', price_raw)
        if pm:
            each_price = float(pm.group(2))

        # Parse conv ratio from pack type e.g. "24/20oz BTL" → 24
        conv = 1.0
        cnm = re.match(r'^(\d+)/', pack_type)
        if cnm:
            conv = float(cnm.group(1))

        # Total each units = case_qty * conv + each_qty
        total_each = case_qty * conv + each_qty

        key = build_key(desc, pack_type)
        if not key:
            continue

        # Accumulate across stands
        counts[key] = counts.get(key, 0.0) + total_each
        if each_price > 0:
            prices[key] = each_price

    return counts, prices

# ── Main UI ───────────────────────────────────────────────────────────────────
st.title("📋 Concessions Database Import")
st.caption("One-shot import: DATABASE sheet → item details · Count sheet → quantities")

db_file    = "/mnt/user-data/uploads/2026_UH__2_.xlsx"
count_file = "/mnt/user-data/uploads/TDECUcombinedvaluesseperatstands.xlsx"

# ── Preview ───────────────────────────────────────────────────────────────────
if st.button("🔍 Preview — Load & Analyze (no writes)", type="secondary"):
    with st.spinner("Reading files…"):
        db_items = load_database_sheet(db_file)
        counts, prices = load_count_sheet(count_file)

    st.success(f"DATABASE sheet: **{len(db_items)}** items")
    st.success(f"Count sheet: **{len(counts)}** unique item keys, "
               f"total value **${sum(counts[k]*prices.get(k,0) for k in counts):,.2f}**")

    # Match analysis
    db_keys   = {i["key"] for i in db_items}
    cnt_keys  = set(counts.keys())
    matched   = db_keys & cnt_keys
    db_only   = db_keys - cnt_keys
    cnt_only  = cnt_keys - db_keys

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("DB items",          len(db_keys))
    c2.metric("Count items",       len(cnt_keys))
    c3.metric("Matched (exact key)",len(matched))
    c4.metric("Count-only (new)",  len(cnt_only))

    if cnt_only:
        with st.expander(f"📋 {len(cnt_only)} items in count but not in DATABASE"):
            st.dataframe(pd.DataFrame(
                [{"Key": k, "Count": counts[k], "Each Price": f"${prices.get(k,0):.2f}"}
                 for k in sorted(cnt_only)]
            ), use_container_width=True, hide_index=True)

    if db_only:
        with st.expander(f"📋 {len(db_only)} DATABASE items with no count"):
            st.dataframe(pd.DataFrame(
                [{"Key": k} for k in sorted(db_only)]
            ), use_container_width=True, hide_index=True)

st.markdown("---")

# ── Step 1: Import DATABASE ───────────────────────────────────────────────────
st.subheader("Step 1 — Import DATABASE sheet (item details + overrides)")
if st.button("▶️ Run Step 1", type="primary", key="step1"):
    db  = get_db()
    with st.spinner("Loading DATABASE sheet…"):
        db_items = load_database_sheet(db_file)

    added = updated = errors = 0
    prog = st.progress(0)
    for i, item in enumerate(db_items):
        try:
            result = db.upsert_item(
                item,
                doc_date=datetime.utcnow().strftime("%Y-%m-%d"),
                source_document="2026_UH__2_.xlsx DATABASE",
                changed_by="database_import",
            )
            if result == "created":
                added += 1
            elif result == "updated":
                updated += 1
        except Exception as e:
            errors += 1
        prog.progress((i+1) / len(db_items))

    prog.empty()
    st.success(f"✅ Step 1 complete — **{added}** added, **{updated}** updated, **{errors}** errors")

st.markdown("---")

# ── Step 2: Import counts ─────────────────────────────────────────────────────
st.subheader("Step 2 — Import count sheet (quantities)")
st.caption("Adds quantity_on_hand to matched items. Creates new items for unmatched count entries.")

if st.button("▶️ Run Step 2", type="primary", key="step2"):
    db = get_db()
    with st.spinner("Loading count sheet…"):
        counts, prices = load_count_sheet(count_file)

    updated = created = errors = 0
    prog = st.progress(0)
    keys = list(counts.keys())

    for i, key in enumerate(keys):
        qty        = counts[key]
        each_price = prices.get(key, 0.0)
        try:
            existing = db.get_item(key)
            if existing:
                from database import get_conn
                with get_conn() as conn:
                    cur = conn.cursor()
                    cur.execute(
                        "UPDATE items SET quantity_on_hand=%s, last_updated=%s WHERE key=%s",
                        (qty, datetime.utcnow(), key)
                    )
                updated += 1
            else:
                # New item from count — parse description and pack from key
                parts = key.split("||")
                desc  = parts[0]
                pack  = parts[1] if len(parts) > 1 else "CASE"
                db.add_item({
                    "key":              key,
                    "description":      desc,
                    "pack_type":        pack,
                    "cost":             each_price,
                    "quantity_on_hand": qty,
                    "cost_center":      "57231",
                    "status_tag":       "📦 Count Import",
                    "record_status":    "active",
                    "conv_ratio":       1.0,
                    "yield":            1.0,
                    "is_chargeable":    True,
                    "last_updated":     datetime.utcnow(),
                    "created_date":     datetime.utcnow(),
                }, changed_by="count_import")
                created += 1
        except Exception as e:
            errors += 1
        prog.progress((i+1) / len(keys))

    prog.empty()
    st.success(
        f"✅ Step 2 complete — "
        f"**{updated}** quantities updated, "
        f"**{created}** new items created, "
        f"**{errors}** errors"
    )

st.markdown("---")

# ── Step 3: Verify ────────────────────────────────────────────────────────────
st.subheader("Step 3 — Verify totals")
st.caption("Compare our DB totals against MyOrders target: $114,371.81")

if st.button("🔍 Check Totals", key="step3"):
    db = get_db()
    with st.spinner("Calculating…"):
        total_value = db.get_inventory_value()
        total_items = db.count_items("active")

    c1, c2, c3 = st.columns(3)
    c1.metric("Total Items",    total_items)
    c2.metric("Our Total Value",f"${total_value:,.2f}")
    c3.metric("MyOrders Target","$114,371.81")

    diff = total_value - 114371.81
    if abs(diff) < 100:
        st.success(f"✅ Within $100 of target — variance: ${diff:+,.2f}")
    elif abs(diff) < 1000:
        st.warning(f"⚠️ Variance: ${diff:+,.2f} — close but check for missing items")
    else:
        st.error(f"❌ Variance: ${diff:+,.2f} — significant gap, investigate")
